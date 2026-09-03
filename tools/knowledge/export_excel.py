"""Paso 3 del pipeline: libro de revision en Excel de todo lo extraido del Estatuto.

Uso:  python tools/knowledge/export_excel.py <corpus.json> <destino.xlsx>

Seis hojas: resumen, preguntas y respuestas curadas, tarifas de ICA, tablas verificadas,
indice de los 726 articulos y pendientes por revisar. Esta pensado para que la entidad
apruebe o corrija el contenido antes de publicarlo.
"""

import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT_NAME = "Arial"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color="1F3864")
BODY_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Etiqueta legible de la procedencia de cada dato, para quien revisa.
CONFIANZA_TEXTO = {
    "curada": "Respuesta redactada y citada al articulado",
    "verificada": "Tabla transcrita por lectura visual y verificada celda por celda",
    "ocr_texto": "Texto reconocido por OCR del PDF escaneado",
    "ocr_geometria": "Tabla reconstruida por posicion de columna (OCR con geometria)",
}


def write_header(sheet, headers, widths, freeze="A2"):
    sheet.append(headers)
    for index, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = freeze
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def style_body(sheet, wrap_columns=(), first_row=2):
    for row in sheet.iter_rows(min_row=first_row, max_row=sheet.max_row):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
            wrap = cell.column_letter in wrap_columns
            cell.alignment = Alignment(vertical="top", wrap_text=wrap)


def refresh_autofilter(sheet):
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"


def sheet_preguntas(book, faqs):
    sheet = book.create_sheet("Preguntas y respuestas")
    write_header(
        sheet,
        ["ID", "Tema", "Pregunta", "Respuesta propuesta", "Palabras clave",
         "Artículos del Estatuto", "Vigencia del dato", "Nota de revisión", "¿Aprobada?", "Corrección de la entidad"],
        [16, 26, 42, 78, 40, 16, 14, 34, 12, 46],
    )

    for faq in faqs:
        sheet.append([
            faq["id"],
            faq.get("tema", ""),
            faq["pregunta"],
            faq["respuesta"],
            ", ".join(faq.get("palabras_clave", [])),
            ", ".join(str(a) for a in faq.get("articulos", [])),
            faq.get("vigencia", "estable"),
            faq.get("nota_revision", ""),
            "",
            "",
        ])

    style_body(sheet, wrap_columns=("C", "D", "E", "H", "J"))

    # Las dos ultimas columnas son las que llena la entidad.
    aprobacion = DataValidation(type="list", formula1='"Sí,No,Ajustar"', allow_blank=True)
    sheet.add_data_validation(aprobacion)
    aprobacion.add(f"I2:I{sheet.max_row}")
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=9).fill = INPUT_FILL
        sheet.cell(row=row, column=10).fill = INPUT_FILL
        # Lo volatil se resalta: son las respuestas que caducan cada ano.
        if sheet.cell(row=row, column=7).value == "volatil":
            sheet.cell(row=row, column=7).fill = WARN_FILL

    refresh_autofilter(sheet)
    return sheet


def sheet_tarifas_ica(book, filas):
    sheet = book.create_sheet("Tarifas ICA")
    write_header(
        sheet,
        ["Código CIIU", "División / grupo", "Actividad económica", "Rango de ingresos brutos anuales",
         "Tarifa (por mil)", "Página del PDF", "Estado", "Corrección de la entidad"],
        [13, 34, 62, 30, 14, 13, 20, 24],
    )

    for fila in filas:
        tarifa = fila.get("tarifa", "")
        sheet.append([
            fila.get("codigo", ""),
            fila.get("division", ""),
            fila.get("actividad", ""),
            fila.get("rango", ""),
            tarifa if tarifa else "sin dato",
            fila.get("pagina_pdf", ""),
            "Extraída" if tarifa else "SIN DATO: verificar en el PDF",
            "",
        ])

    style_body(sheet, wrap_columns=("B", "C", "D", "G"))
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=8).fill = INPUT_FILL
        if not str(sheet.cell(row=row, column=5).value or "").strip() or sheet.cell(row=row, column=5).value == "sin dato":
            for col in range(1, 8):
                sheet.cell(row=row, column=col).fill = WARN_FILL

    refresh_autofilter(sheet)
    return sheet


def sheet_tablas(book, tablas):
    sheet = book.create_sheet("Tablas verificadas")
    write_header(
        sheet,
        ["Tabla", "Artículo", "Página del PDF", "Unidad", "Columna", "Fila", "Valor"],
        [40, 10, 14, 24, 26, 32, 14],
    )

    for tabla in tablas:
        columnas = tabla["columnas"]
        for fila in tabla["filas"]:
            etiqueta = " / ".join(str(v) for v in fila[: max(1, len(columnas) - len(fila) + 1)][:2] if v not in ("", "-"))
            for indice, valor in enumerate(fila):
                if indice < len(columnas) and columnas[indice].lower().startswith(("rango", "destinacion", "estratificacion")):
                    continue
                sheet.append([
                    tabla["titulo"],
                    tabla["articulo"],
                    tabla["pagina_pdf"],
                    tabla["unidad"],
                    columnas[indice] if indice < len(columnas) else f"columna {indice + 1}",
                    etiqueta,
                    valor,
                ])

    style_body(sheet, wrap_columns=("A", "D", "F"))
    refresh_autofilter(sheet)
    return sheet


def sheet_indice(book, chunks):
    sheet = book.create_sheet("Índice de artículos")
    write_header(
        sheet,
        ["Artículo", "Epígrafe", "Tema", "Libro", "Capítulo", "Página del PDF", "Fragmentos", "Caracteres"],
        [10, 56, 30, 30, 40, 14, 12, 12],
    )

    por_articulo = {}
    for chunk in chunks:
        if chunk["tipo"] != "prosa":
            continue
        entry = por_articulo.setdefault(
            chunk["articulo"],
            {"epigrafe": chunk["epigrafe"], "tema": chunk["tema"], "libro": chunk["libro"],
             "capitulo": chunk["capitulo"], "pagina": chunk["pagina_pdf"], "fragmentos": 0, "chars": 0},
        )
        entry["fragmentos"] += 1
        entry["chars"] += len(chunk["texto"])
        if not entry["epigrafe"] and chunk["epigrafe"]:
            entry["epigrafe"] = chunk["epigrafe"]

    for numero in sorted(por_articulo):
        entry = por_articulo[numero]
        sheet.append([
            numero, entry["epigrafe"], entry["tema"], entry["libro"], entry["capitulo"],
            entry["pagina"], entry["fragmentos"], entry["chars"],
        ])

    style_body(sheet, wrap_columns=("B", "C", "D", "E"))
    for row in range(2, sheet.max_row + 1):
        if not str(sheet.cell(row=row, column=2).value or "").strip():
            sheet.cell(row=row, column=2).value = "(sin epígrafe legible)"
            sheet.cell(row=row, column=2).fill = WARN_FILL

    refresh_autofilter(sheet)
    return sheet


def sheet_revisar(book, informe, ica):
    sheet = book.create_sheet("Por revisar")
    write_header(
        sheet,
        ["Tipo", "Detalle", "Página del PDF", "Qué hay que hacer", "Resuelto por la entidad"],
        [30, 78, 14, 52, 26],
    )

    for aviso in ica.get("avisos", []):
        if aviso["tipo"] == "tarifa_ausente":
            sheet.append([
                "Tarifa de ICA sin dato",
                f"Código {aviso.get('codigo','')} - {str(aviso.get('actividad',''))[:120]} - {aviso.get('rango','')}",
                aviso.get("pagina_pdf", ""),
                "El OCR no reconoció el dígito. Transcribir la tarifa desde el PDF.",
                "",
            ])
        else:
            sheet.append([
                f"Tarifa de ICA inconsistente ({aviso['tipo']})",
                f"Código {aviso.get('codigo','')} - {str(aviso.get('actividad',''))[:120]} - valores {aviso.get('valores', aviso.get('valor',''))}",
                aviso.get("pagina_pdf", ""),
                "La tarifa no crece con el rango de ingresos. Verificar en el PDF.",
                "",
            ])

    for correccion in ica.get("correcciones", []):
        sheet.append([
            "Coma decimal repuesta automáticamente",
            f"Código {correccion.get('codigo','')} - {str(correccion.get('actividad',''))[:120]} - leído \"{correccion['leido']}\", corregido a \"{correccion['corregido']}\"",
            correccion.get("pagina_pdf", ""),
            "Supera el techo legal de 12 por mil, así que la coma se repuso. Confirmar.",
            "",
        ])

    for descartada in informe.get("cabeceras_descartadas", []):
        sheet.append([
            "Cabecera de artículo descartada",
            f"Se leyó \"ARTÍCULO {descartada['numero']}\" fuera de la secuencia: {descartada['texto'][:100]}",
            descartada.get("pagina_pdf", ""),
            "Es una referencia a otra norma, no un artículo del Acuerdo. Confirmar.",
            "",
        ])

    for numero in informe.get("numeros_sin_detectar", []):
        sheet.append([
            "Artículo no detectado",
            f"No se encontró la cabecera del artículo {numero}",
            "",
            "Revisar la página correspondiente del PDF.",
            "",
        ])

    style_body(sheet, wrap_columns=("A", "B", "D"))
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=5).fill = INPUT_FILL
    refresh_autofilter(sheet)
    return sheet


def sheet_resumen(book, corpus, hojas):
    sheet = book["Resumen"]
    sheet.column_dimensions["A"].width = 46
    sheet.column_dimensions["B"].width = 62

    fuente = corpus.get("fuente", {})
    filas = [
        ("Base de conocimiento del asistente virtual", ""),
        ("", ""),
        ("Documento de origen", fuente.get("documento", "")),
        ("Fecha del Acuerdo", fuente.get("fecha", "")),
        ("Páginas del PDF", fuente.get("paginas", "")),
        ("Método de extracción", fuente.get("extraccion", "")),
        ("", ""),
        ("Contenido de este libro", ""),
    ]
    for etiqueta, valor in filas:
        sheet.append([etiqueta, valor])

    conteos = [
        ("Preguntas y respuestas propuestas", f"=COUNTA('Preguntas y respuestas'!A2:A100000)"),
        ("Filas de tarifas de ICA extraídas", f"=COUNTA('Tarifas ICA'!A2:A100000)"),
        ("Tarifas de ICA sin dato", f"=COUNTIF('Tarifas ICA'!E2:E100000,\"sin dato\")"),
        ("Valores de tablas verificadas a mano", f"=COUNTA('Tablas verificadas'!A2:A100000)"),
        ("Artículos del Estatuto en el corpus", f"=COUNTA('Índice de artículos'!A2:A100000)"),
        ("Pendientes por revisar", f"=COUNTA('Por revisar'!A2:A100000)"),
    ]
    for etiqueta, formula in conteos:
        sheet.append([etiqueta, formula])

    sheet.append(["", ""])
    sheet.append(["Cómo leer la procedencia de cada dato", ""])
    for clave, texto in CONFIANZA_TEXTO.items():
        sheet.append([clave, texto])

    sheet.append(["", ""])
    sheet.append(["Qué debe llenar la entidad", ""])
    leyenda = [
        ("Celdas con fondo amarillo", "Son las únicas que hay que escribir. El resto es contenido extraído."),
        ("Hoja Preguntas y respuestas", "Columna «¿Aprobada?» (valores: Sí, No, Ajustar) y «Corrección de la entidad»."),
        ("Hoja Tarifas ICA", "Columna «Corrección de la entidad» en las filas marcadas SIN DATO."),
        ("Hoja Por revisar", "Columna «Resuelto por la entidad»."),
        ("Filas con fondo naranja", "Requieren atención: dato faltante, inconsistente o sin epígrafe legible."),
        ("Columna «Vigencia del dato»", "«volatil» significa que caduca cada año (calendario, descuentos, valor de la UVT)."),
    ]
    for etiqueta, texto in leyenda:
        sheet.append([etiqueta, texto])

    sheet.append(["", ""])
    sheet.append(["Advertencia sobre las cifras", ""])
    sheet.append([
        "Origen escaneado",
        "El PDF no tiene capa de texto: todo se obtuvo por OCR. Las tarifas de las tablas "
        "matriciales se transcribieron a mano porque el OCR parte las celdas. Antes de "
        "publicar una cifra, confírmela contra el PDF oficial.",
    ])

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet["A1"].font = TITLE_FONT
    for etiqueta in ("Contenido de este libro", "Cómo leer la procedencia de cada dato",
                     "Qué debe llenar la entidad", "Advertencia sobre las cifras"):
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == etiqueta:
                sheet.cell(row=row, column=1).font = Font(name=FONT_NAME, size=11, bold=True, color="1F3864")
    return sheet


def main(corpus_path, dest_path):
    corpus = json.load(open(corpus_path, encoding="utf-8"))
    here = os.path.dirname(os.path.abspath(__file__))
    faqs = json.load(open(os.path.join(here, "manual", "preguntas.json"), encoding="utf-8"))["preguntas"]
    tablas = json.load(open(os.path.join(here, "manual", "tablas_verificadas.json"), encoding="utf-8"))["tablas"]
    ica = json.load(open(os.path.join(here, "out", "tarifas_ica.json"), encoding="utf-8"))
    informe = json.load(open(os.path.join(here, "out", "informe_extraccion.json"), encoding="utf-8"))

    book = Workbook()
    book.active.title = "Resumen"

    hojas = {
        "preguntas": sheet_preguntas(book, faqs),
        "ica": sheet_tarifas_ica(book, ica["filas"]),
        "tablas": sheet_tablas(book, tablas),
        "indice": sheet_indice(book, corpus["chunks"]),
        "revisar": sheet_revisar(book, informe, ica),
    }
    sheet_resumen(book, corpus, hojas)

    os.makedirs(os.path.dirname(os.path.abspath(dest_path)), exist_ok=True)
    book.save(dest_path)

    print(f"libro            : {dest_path} ({os.path.getsize(dest_path)/1e6:.2f} MB)")
    for nombre, hoja in hojas.items():
        print(f"  {nombre:10} {hoja.max_row - 1} filas")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(2)
    sys.exit(main(sys.argv[1], sys.argv[2]))
